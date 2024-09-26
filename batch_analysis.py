import os
import argparse
from concurrent.futures import ProcessPoolExecutor
from functions.helper import align_row
from pysanger import visualize
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import datetime
    
def batch_alignments(setup="seq_results/Sequencing_setup.xlsx", input_folder="seq_results", template_folder="templates"):
    """
    Aligns sequencing files to template files in parallel using multiple CPU cores.

    Args:
        setup (str): The path to the sequencing setup file (default: "seq_results/Sequencing_setup.xlsx").
        input_folder (str): The path to the folder containing the sequencing files (default: "seq_results").
        template_folder (str): The path to the folder containing the template files (default: "templates").

    Returns:
        dict: A dictionary containing the alignments, where the keys are Sequencing_IDs and the values are alignment results.
    """

    # Read the sequencing setup file
    setup_file = pd.read_excel(setup, engine='openpyxl', sheet_name='Sheet1', header=0)
    
    required_columns = ['Position', 'Sequencing_ID', 'Template', 'Primer', 'strand']
    assert all(column in setup_file.columns for column in required_columns), "The sequencing setup file must contain all the required columns: Position, Sequencing_ID, Template, Primer, strand"
    assert setup_file['Sequencing_ID'].is_unique, "The Sequencing_ID column must contain unique values"
    assert all(strand == -1 or strand == 1 for strand in setup_file['strand']), "The strand column must contain only -1 or 1"

    # Create a list of all the sequencing files
    files = [f for f in os.listdir(input_folder) if f.endswith(".ab1")]
    
    sequencing_files = []
    for f in files:
        try:
            sequencing_file = f.split(".", maxsplit=1)[0].split("-", maxsplit=1)[0]
            sequencing_files.append(sequencing_file)
        except IndexError:
            print(f"Filename '{f}' cannot be split by the expected '-' character into Sequencing_ID and Primer. `{f}` skipped.")
            continue
    sequencing_values = np.unique(setup_file['Sequencing_ID'].values)
    # Identify sequencing files that are not in the sequencing ID list
    missing_sequencing_files = [seq for seq in sequencing_values if seq not in sequencing_files]

    # If there are missing sequencing files, print them and raise an assertion error
    if missing_sequencing_files:
        print(f"Missing sequencing files: {missing_sequencing_files}")
        assert False, "The Sequencing_ID column in the sequencing setup file must contain all the sequencing IDs in the input folder"

    primer_files = []
    for f in files:
        try:
            primer_file = f.split(".", maxsplit=1)[0].split("-", maxsplit=1)[1]
            primer_files.append(primer_file)
        except IndexError:
            print(f"Filename '{f}' cannot be split by the expected '-' character into Sequencing_ID and Primer. `{f}` skipped.")
            continue
    primer_values = np.unique(setup_file['Primer'].values)
    # Identify primers that are not in the primer list
    missing_primers = [primer for primer in primer_values if primer not in primer_files]

    # If there are missing primers, print them and raise an assertion error
    if missing_primers:
        print(f"Missing primers: {missing_primers}")
        assert False, "The Primer column in the sequencing setup file must contain all the primers in the input folder"


    # Create a list of all the template files
    templates = [f for f in os.listdir(template_folder) if f.endswith((".dna", ".gb", ".xdna"))]
    template_ids = [f.split("_", maxsplit=1)[0] for f in templates] # Extract the template ID from the file name by splitting filenames by '_'
    template_names = [f.split("_", maxsplit=1)[1] for f in templates] # Extract the template names from the file name by splitting filenames by '_'
    template_values = np.unique(setup_file['Template'].values)
    # Identify template files that are not in the template ID list
    missing_template_ids = [template for template in template_ids if template not in template_values]

    # If there are missing template IDs, print them and raise an assertion error
    if missing_template_ids:
        print(f"Missing template IDs: {missing_template_ids}")
        assert False, "The Template column in the sequencing setup file must contain all the template IDs in the template folder"

    strand_values = setup_file['strand'].values
    
    # Create a data frame for sequencing reactions
    sequencing_dict = {seq_id: file for seq_id, file in zip(sequencing_files, files)}
    template_dict = {temp_id: file for temp_id, file in zip(template_ids, templates)}
    strand_dict = {seq_id: strand for seq_id, strand in zip(sequencing_values, strand_values)}
    sequencing_reactions = setup_file[['Sequencing_ID', 'Template']].copy()

    # Map the sequencing and template files to the DataFrame
    sequencing_reactions['Seq_File'] = sequencing_reactions['Sequencing_ID'].map(sequencing_dict)
    sequencing_reactions['Template_File'] = sequencing_reactions['Template'].map(template_dict)
    sequencing_reactions['strand'] = sequencing_reactions['Sequencing_ID'].map(strand_dict)
    sequencing_reactions = sequencing_reactions[['Seq_File', 'Template_File', 'strand']]
    # Create a dictionary to store the alignments
    alignments = {}

    # Use ProcessPoolExecutor to parallelize the alignment process
    num_cpus = os.cpu_count()
    max_workers = max(1, num_cpus - 2)  # Ensure at least 1 worker
    print("Using", max_workers, "of", num_cpus, "available CPUs for parallel alignment.")

    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        try:
            futures = {
                executor.submit(align_row, row, input_folder, template_folder): row
                for index, row in sequencing_reactions.iterrows()
            }
        except Exception as e:
            print(f"Error processing {row['Seq_File']}: {e}")
        for future in futures:
            try:
                sample_name, template_name, alignment_result = future.result()
                name = sample_name.split("-")[0]
                alignments[name] = alignment_result
            except Exception as e:
                print(f"Error processing {futures[future]}: {e}")
    
    return alignments, setup_file

def summarise_results(batch_alignments, output_folder="output"):
    """
    Summarize the alignment results for each sample and write an excel file with the summary to an output folder.

    Args:
        batch_alignments (dict): A dictionary containing the alignments, where the keys are Sequencing_IDs and the values are alignment results.
        output_folder (str, optional): The folder where the alignment summary will be saved. Defaults to "output".

    Returns:
        pandas.DataFrame: A DataFrame containing the alignment summary.
    """
    alignments, setup = batch_alignments
    summaries = {}
    for sample, alignment in alignments.items():
        _abidata, asubject, subject, atemplate, sequence, matches, ss, se, ts, te, avalues, tvalues, gvalues, cvalues, name, seq_filename, strand = alignment
        if ts != 0:
            ss = ts
        elif ts == 0:
            ts = ss
        if te != 0:
            se = te
        elif te == 0:
            te = se
        tmatches = matches[ts:-te]
        
        match_count = matches.count(1)
        mismatch_count = matches.count(-1)
        not_aligned_count = matches.count(0)
        if strand == 1:
            if len(tmatches) <800: 
                match_count_stringent = tmatches[25:].count(1)
                mismatch_count_stringent = tmatches[25:].count(-1)
                not_aligned_count_stringent = tmatches[50:].count(0)
            else:
                match_count_stringent = tmatches[25:800].count(1)
                mismatch_count_stringent = tmatches[25:800].count(-1)
                not_aligned_count_stringent = tmatches[25:800].count(0)
        elif strand == -1:
            if len(tmatches) <800: 
                match_count_stringent = tmatches[:-25].count(1)
                mismatch_count_stringent = tmatches[:-25].count(-1)
                not_aligned_count_stringent = tmatches[:-25].count(0)
            else:
                match_count_stringent = tmatches[len(tmatches)-800:-25].count(1)
                mismatch_count_stringent = tmatches[len(tmatches)-800:-25].count(-1)
                not_aligned_count_stringent = tmatches[len(tmatches)-800:-25].count(0)
        

        summary = {
            "Sequencing_ID": sample,
            "Template_file": name,
            "Template_length": len(sequence),
            "Sequencing_file": seq_filename,
            "Sequencing_length": len(subject),
            "Alignment_length": len(matches),
            "Alignment_start": ts,
            "Alignment_end": len(sequence)-te,
            "Matches": match_count,
            "Mismatches": mismatch_count,
            "Not_aligned": not_aligned_count,
            "Stringent_start": 25,
            "Stringent_end": 800,
            "Stringend_Matches": match_count_stringent,
            "Stringent_Mismatches": mismatch_count_stringent,
            "Stringent_Not_aligned": not_aligned_count_stringent
        }
        summaries[sample] = summary
    
    alignment_df = pd.DataFrame.from_dict(summaries, orient='index')
    merged_df = pd.merge(setup, alignment_df, on='Sequencing_ID')

    # Ensure the output folder exists
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    output_file = f"{output_folder}/{datetime.datetime.now().strftime('%Y-%m-%d')}_alignment_summary.xlsx"
    merged_df.to_excel(output_file, index=False)
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    column_index = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == "Stringent_Mismatches":
            column_index = col[0].column
            break

    if column_index:
        for cell in ws[get_column_letter(column_index)][1:]:
            if cell.value and cell.value > 0: 
                cell.font = Font(bold=True, color="FF0000")
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    wb.save(output_file)
    
    return merged_df
                
def batch_visualization(batch_alignments, region="aligned", output_folder="output"):
    """
    Visualizes batch alignments and saves the alignment chromatograms as PDF files.

    Args:
        batch_alignments (tuple): A tuple containing the alignments and setup information.
        region (str, optional): The region of the alignment to visualize. Defaults to "aligned".
        output_folder (str, optional): The folder where the alignment chromatograms will be saved. Defaults to "output".

    Returns:
        None
    """
    # Ensure the output folder exists
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    alignments, setup = batch_alignments
    for sample, alignment in alignments.items():
        fig = visualize(alignment, region=region, fig_width=100, fontsize=4)
        print(f"Saving alignment chromatogram for {sample} to {output_folder}/{sample}_alignment.pdf")
        fig.savefig(f"{output_folder}/{datetime.datetime.now().strftime('%Y-%m-%d')}_{sample}_alignment.pdf")
    
def main(setup="seq_results/Sequencing_setup.xlsx", input_folder="seq_results", template_folder="templates", region="aligned", output_folder="output"):
    """
    Main function to align sequencing files to template files, summarize the alignment results, and visualize the alignments.

    Args:
        setup (str): The path to the sequencing setup file (default: "seq_results/Sequencing_setup.xlsx").
        input_folder (str): The path to the folder containing the sequencing files (default: "seq_results").
        template_folder (str): The path to the folder containing the template files (default: "templates").
        output_folder (str): The folder where the alignment summary and chromatograms will be saved (default: "output").
        region (str): The region of the alignment to visualize (default: "aligned").

    Returns:
        None
    """
    batch = batch_alignments(setup=setup, input_folder=input_folder, template_folder=template_folder)
    summary = summarise_results(batch, output_folder=output_folder)
    batch_visualization(batch, region=region, output_folder=output_folder)

if __name__ == "__main__":

    parser = argparse.ArgumentParser(description="Run batch alignment analysis.")
    parser.add_argument("--setup", required=True, help="Path to the setup file.")
    parser.add_argument("--input_folder", required=True, help="Path to the input folder.")
    parser.add_argument("--template_folder", required=True, help="Path to the template folder.")
    parser.add_argument("--region", required=True, help="Region for visualization.")
    parser.add_argument("--output_folder", required=True, help="Path to the output folder.")
    
    args = parser.parse_args()
    
    main(args.setup, args.input_folder, args.template_folder, args.region, args.output_folder)
